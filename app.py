import streamlit as st
import os
import json
import shutil
import re
import datetime
import zipfile
import io
import csv
import pandas as pd
import threading
import time

def start_session_monitor():
    """
    Spawns a background thread to monitor active Streamlit websocket connections.
    If no browser tabs are active for 10 seconds, shuts down the python process automatically.
    """
    if any(t.name == "SessionMonitorThread" for t in threading.enumerate()):
        return

    def monitor_sessions():
        time.sleep(15)  # Wait for startup and initial connection
        inactive_streak = 0
        while True:
            time.sleep(2)
            try:
                from streamlit.runtime import runtime
                rt = runtime.get_instance()
                if rt:
                    active_sessions = len(rt._session_info_by_id)
                    if active_sessions == 0:
                        inactive_streak += 2
                        if inactive_streak >= 10:  # No tab active for 10 seconds
                            os._exit(0)
                    else:
                        inactive_streak = 0
            except Exception:
                pass

    t = threading.Thread(target=monitor_sessions, name="SessionMonitorThread", daemon=True)
    t.start()

start_session_monitor()

# Use st.dialog if available, else fallback
if hasattr(st, "dialog"):
    @st.dialog("✨ ดาวน์โหลดเสร็จสมบูรณ์")
    def show_success_modal(count, total_cnt):
        st.balloons()
        st.markdown(f"""
        <div style='text-align: center; padding: 10px 0;'>
            <h2 style='color: #10B981; margin-bottom: 10px;'>🎉 ดาวน์โหลดเสร็จสมบูรณ์!</h2>
            <p style='font-size: 1.1em;'>ระบบได้คัดแยกไฟล์และลงตาราง Excel เรียบร้อยแล้ว</p>
            <div style='background-color: #F3F4F6; padding: 15px; border-radius: 8px; margin: 15px 0;'>
                <span style='font-size: 1.25em; font-weight: 600;'>🏥 นำเข้าสำเร็จทั้งหมด: <span style='color:#10B981;'>{count}</span> จาก {total_cnt} โรงพยาบาล</span>
            </div>
            <p style='color: #6B7280; font-size: 0.9em;'>ไฟล์แนบจะถูกเก็บแยกในโฟลเดอร์ 1. นำเข้าใหม่ยังไม่ได้เช็ค</p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("ตกลง (OK)", use_container_width=True):
            st.rerun()
elif hasattr(st, "experimental_dialog"):
    @st.experimental_dialog("✨ ดาวน์โหลดเสร็จสมบูรณ์")
    def show_success_modal(count, total_cnt):
        st.balloons()
        st.markdown(f"""
        <div style='text-align: center; padding: 10px 0;'>
            <h2 style='color: #10B981; margin-bottom: 10px;'>🎉 ดาวน์โหลดเสร็จสมบูรณ์!</h2>
            <p style='font-size: 1.1em;'>ระบบได้คัดแยกไฟล์และลงตาราง Excel เรียบร้อยแล้ว</p>
            <div style='background-color: #F3F4F6; padding: 15px; border-radius: 8px; margin: 15px 0;'>
                <span style='font-size: 1.25em; font-weight: 600;'>🏥 นำเข้าสำเร็จทั้งหมด: <span style='color:#10B981;'>{count}</span> จาก {total_cnt} โรงพยาบาล</span>
            </div>
            <p style='color: #6B7280; font-size: 0.9em;'>ไฟล์แนบจะถูกเก็บแยกในโฟลเดอร์ 1. นำเข้าใหม่ยังไม่ได้เช็ค</p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("ตกลง (OK)", use_container_width=True):
            st.rerun()
else:
    def show_success_modal(count, total_cnt):
        st.success(f"🎉 ดาวน์โหลดเสร็จสมบูรณ์! นำเข้าสำเร็จทั้งหมด {count} จาก {total_cnt} โรงพยาบาล")

# Import local handlers (TK-independent)
from email_handler import (
    fetch_emails_by_range, 
    extract_text_from_pdf, 
    parse_email_date_to_yyyy_mm_dd, 
    check_folder_completeness
)
from gemini_handler import analyze_email_with_gemini
from excel_handler import append_to_excel

# Constants
CONFIG_FILE = "config.json"
ASSESSORS_FILE = "assessors.json"
HOSPITAL_REGISTRY_FILE = "hospitals_registry.json"
USAGE_LOG_FILE = "usage_logs.csv"
USER_REGISTRY_FILE = "users.json"

import hashlib

def hash_password(password, salt=None):
    if salt is None:
        salt = os.urandom(16).hex()
    hashed = hashlib.sha256((password + salt).encode('utf-8')).hexdigest()
    return hashed, salt

def load_users():
    if os.path.exists(USER_REGISTRY_FILE):
        try:
            with open(USER_REGISTRY_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def save_users(users):
    try:
        with open(USER_REGISTRY_FILE, "w", encoding="utf-8") as f:
            json.dump(users, f, ensure_ascii=False, indent=4)
        return True
    except Exception:
        return False

def init_default_user():
    users = load_users()
    if "khem" not in users:
        hashed, salt = hash_password("123456")
        users["khem"] = {
            "username": "khem",
            "password_hash": hashed,
            "salt": salt,
            "display_name": "พี่เข้ม",
            "fullname": "ทนพ.สุทัศน์ บุญยงค์",
            "role": "อนุกรรมการตรวจประเมิน LA สภาเทคนิคการแพทย์ 2569"
        }
        save_users(users)

init_default_user()

def render_login_page():
    # Centering Layout using Streamlit columns
    _, col_login, _ = st.columns([1, 1.8, 1])
    
    with col_login:
        st.write("<div style='height: 40px;'></div>", unsafe_allow_html=True)
        # Large Branding Title
        st.markdown("""
        <div style='text-align: center; margin-bottom: 25px;'>
            <div style='background: linear-gradient(135deg, #38BDF8 0%, #EC4899 100%); -webkit-background-clip: text; -webkit-text-fill-color: transparent; font-size: 28px; font-weight: 900; line-height: 1.3;'>Laboratory Accreditation Management System (LAMS)</div>
            <div style='color: #94A3B8; font-size: 13px; font-style: italic; margin-top: 8px;'>Integrated Quality, Evidence & Assessment Management Platform</div>
        </div>
        """, unsafe_allow_html=True)
        
        tab_login, tab_register = st.tabs(["🔑 เข้าสู่ระบบ (Login)", "📝 สมัครสมาชิกใหม่ (Sign Up)"])
        
        with tab_login:
            st.markdown("<div style='height: 15px;'></div>", unsafe_allow_html=True)
            with st.form("form_login", border=False):
                u_name = st.text_input("ชื่อผู้ใช้งาน (Username):", key="login_username")
                u_pwd = st.text_input("รหัสผ่าน (Password):", type="password", key="login_password")
                
                st.markdown("<div style='height: 15px;'></div>", unsafe_allow_html=True)
                login_submitted = st.form_submit_button("🔑 ลงชื่อเข้าใช้งาน", use_container_width=True, type="primary")
                if login_submitted:
                    if not u_name or not u_pwd:
                        st.error("กรุณาระบุชื่อผู้ใช้งานและรหัสผ่าน!")
                    else:
                        users = load_users()
                        user = users.get(u_name.strip().lower())
                        if user:
                            hashed, _ = hash_password(u_pwd, user["salt"])
                            if hashed == user["password_hash"]:
                                st.session_state.logged_in = True
                                st.session_state.user_info = user
                                st.session_state.operator_name = user["fullname"]
                                st.success("เข้าสู่ระบบสำเร็จ!")
                                time.sleep(0.5)
                                st.rerun()
                            else:
                                st.error("รหัสผ่านไม่ถูกต้อง!")
                        else:
                            st.error("ไม่พบชื่อผู้ใช้งานนี้ในระบบ!")
                        
        with tab_register:
            st.markdown("<div style='height: 15px;'></div>", unsafe_allow_html=True)
            with st.form("form_register", border=False):
                reg_username = st.text_input("ชื่อผู้ใช้งานที่ต้องการสร้าง (Username):", key="reg_username", help="ใช้ในการเข้าสู่ระบบภาษาอังกฤษตัวเล็กเท่านั้น")
                reg_pwd = st.text_input("รหัสผ่าน (Password):", type="password", key="reg_pwd")
                reg_pwd_conf = st.text_input("ยืนยันรหัสผ่าน (Confirm Password):", type="password", key="reg_pwd_conf")
                reg_disp = st.text_input("ชื่อเล่น / ชื่อย่อแสดงผล (e.g. พี่เข้ม, ปองพล):", key="reg_disp")
                reg_full = st.text_input("ชื่อ-นามสกุลจริง (e.g. ทนพ.สุทัศน์ บุญยงค์):", key="reg_full")
                reg_role = st.text_input("ตำแหน่ง / บทบาท (e.g. อนุกรรมการตรวจประเมิน LA...):", key="reg_role")
                
                st.markdown("<div style='height: 15px;'></div>", unsafe_allow_html=True)
                reg_submitted = st.form_submit_button("📝 ลงทะเบียนสมาชิกใหม่", use_container_width=True, type="primary")
                if reg_submitted:
                    username_clean = reg_username.strip().lower()
                    if not username_clean or not reg_pwd or not reg_disp or not reg_full or not reg_role:
                        st.error("กรุณากรอกข้อมูลให้ครบทุกช่อง!")
                    elif not re.match(r"^[a-zA-Z0-9_\-]+$", username_clean):
                        st.error("ชื่อผู้ใช้งานต้องเป็นตัวอักษรภาษาอังกฤษหรือตัวเลขเท่านั้น!")
                    elif reg_pwd != reg_pwd_conf:
                        st.error("รหัสผ่านและการยืนยันรหัสผ่านไม่ตรงกัน!")
                    else:
                        users = load_users()
                        if username_clean in users:
                            st.error("ขออภัย! ชื่อผู้ใช้งานนี้ถูกใช้ไปแล้วในระบบ")
                        else:
                            hashed, salt = hash_password(reg_pwd)
                            users[username_clean] = {
                                "username": username_clean,
                                "password_hash": hashed,
                                "salt": salt,
                                "display_name": reg_disp.strip(),
                                "fullname": reg_full.strip(),
                                "role": reg_role.strip()
                            }
                            if save_users(users):
                                st.success("สมัครสมาชิกสำเร็จ! กรุณากดไปที่แท็บ 'เข้าสู่ระบบ' เพื่อใช้งาน")
                            else:
                                st.error("เกิดข้อผิดพลาดในการบันทึกบัญชีสมาชิก!")

# Page Configuration
st.set_page_config(
    page_title="การบริหารจัดการเอกสารตรวจประเมิน LA",
    page_icon="🏥",
    layout="wide"
)

if "theme_mode" not in st.session_state:
    st.session_state.theme_mode = "☀️ Bright / Light Mode"

is_dark_theme = (st.session_state.theme_mode == "🌙 Dark Mode")

# Theme Palette Tokens (Green, White, Purple)
if is_dark_theme:
    css_bg_main = "#0B0F19"
    css_bg_sidebar = "#151E2E"
    css_text_primary = "#F8FAFC"
    css_text_secondary = "#94A3B8"
    css_card_bg = "rgba(21, 30, 46, 0.85)"
    css_card_border = "rgba(16, 185, 129, 0.25)"
    css_widget_bg = "#151E2E"
    css_widget_border = "rgba(139, 92, 246, 0.4)"
    css_accent_green = "#10B981"
    css_accent_purple = "#8B5CF6"
    css_grad_title = "linear-gradient(135deg, #34D399 0%, #C084FC 100%)"
    css_btn_sec_bg = "linear-gradient(135deg, #151E2E 0%, #1E293B 100%)"
    css_popover_bg = "#151E2E"
else:
    css_bg_main = "#F8FAFC"
    css_bg_sidebar = "#F1F5F9"
    css_text_primary = "#0F172A"
    css_text_secondary = "#475569"
    css_card_bg = "#FFFFFF"
    css_card_border = "rgba(16, 185, 129, 0.3)"
    css_widget_bg = "#FFFFFF"
    css_widget_border = "rgba(124, 58, 237, 0.35)"
    css_accent_green = "#059669"
    css_accent_purple = "#7C3AED"
    css_grad_title = "linear-gradient(135deg, #059669 0%, #7C3AED 100%)"
    css_btn_sec_bg = "linear-gradient(135deg, #FFFFFF 0%, #E2E8F0 100%)"
    css_popover_bg = "#FFFFFF"

# Custom Dynamic Styling
st.markdown(f"""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Prompt:wght@300;400;500;600;700&family=Sarabun:wght@300;400;500;700&display=swap');
    
    /* Hide default Streamlit top decoration line */
    div[data-testid="stDecoration"] {{
        display: none !important;
        visibility: hidden !important;
        height: 0px !important;
    }}
    
    header[data-testid="stHeader"] {{
        background-color: transparent !important;
        background: transparent !important;
        box-shadow: none !important;
        border: none !important;
        pointer-events: none !important;
    }}
    
    header[data-testid="stHeader"] button, 
    header[data-testid="stHeader"] [data-testid="collapsedControl"],
    [data-testid="collapsedControl"] {{
        pointer-events: auto !important;
    }}
    
    div[data-testid="stMainMenu"], 
    button[aria-label="View Menu"],
    [data-testid="stHeaderActionElements"],
    .stDeployButton,
    button[data-testid="stDeployButton"] {{
        display: none !important;
        visibility: hidden !important;
        width: 0px !important;
        height: 0px !important;
    }}
    
    /* Compact Streamlit UI spacing and expanded width */
    .block-container {{
        max-width: 95% !important;
        padding-top: 1.5rem !important;
        padding-bottom: 1rem !important;
        padding-left: 1.5rem !important;
        padding-right: 1.5rem !important;
    }}
    
    /* Global Font & Dynamic Theme Background */
    html, body, .stApp, [data-testid="stAppViewContainer"], [data-testid="stMain"], .main {{
        font-family: 'Prompt', 'Sarabun', sans-serif !important;
        background-color: {css_bg_main} !important;
        color: {css_text_primary} !important;
    }}

    [data-testid="stSidebar"], [data-testid="stSidebarContent"], [data-testid="stSidebarNav"] {{
        background-color: {css_bg_sidebar} !important;
    }}
    
    h1, h2, h3, h4, h5, h6, input, select, textarea, div.stMarkdown, div[data-testid="stMarkdownContainer"] {{
        font-family: 'Prompt', 'Sarabun', sans-serif !important;
        color: {css_text_primary} !important;
    }}
    
    h1 {{ font-size: 1.6em !important; font-weight: 700 !important; color: {css_text_primary} !important; }}
    h2 {{ font-size: 1.35em !important; font-weight: 600 !important; color: {css_text_primary} !important; }}
    h3 {{ font-size: 1.15em !important; font-weight: 500 !important; color: {css_text_primary} !important; }}
    
    p:not([class*="st-ae"]):not([class*="st-af"]):not([class*="st-ag"]):not([class*="st-ah"]),
    label:not([class*="st-ae"]):not([class*="st-af"]):not([class*="st-ag"]):not([class*="st-ah"]),
    button:not([class*="st-ae"]):not([class*="st-af"]):not([class*="st-ag"]):not([class*="st-ah"]) {{
        font-family: 'Prompt', 'Sarabun', sans-serif !important;
        font-size: 13px !important;
        color: {css_text_primary} !important;
    }}
    
    [data-testid="stIcon"], [class*="icon"], [class*="Icon"], [class*="symbol"], [class*="Symbol"], svg, i {{
        font-family: "Material Icons", "Material Symbols Outlined", "Material Symbols Rounded", inherit !important;
    }}
    
    /* Form Widget Labels as High-Contrast Colored Blocks */
    div[data-testid="stWidgetLabel"] p, 
    div[data-testid="stWidgetLabel"] label {{
        display: inline-block !important;
        background-color: {css_widget_bg} !important;
        color: {css_text_primary} !important;
        padding: 6px 12px !important;
        border-radius: 6px !important;
        font-size: 13.5px !important;
        font-weight: 500 !important;
        margin-bottom: 8px !important;
        border-left: 4px solid {css_accent_green} !important; /* Emerald Green Accent */
        box-shadow: 0 2px 4px rgba(0,0,0,0.1) !important;
        border-top: 1px solid {css_widget_border} !important;
        border-right: 1px solid {css_widget_border} !important;
        border-bottom: 1px solid {css_widget_border} !important;
    }}
    
    /* Form Inputs and Selectboxes - High Contrast Green & Purple Theme */
    div[data-baseweb="select"],
    div[data-baseweb="select"] > div,
    div[data-baseweb="select"] > div > div,
    div[data-baseweb="select"] span,
    div[data-baseweb="select"] p,
    div[data-baseweb="select"] input,
    div[data-baseweb="input"],
    div[data-baseweb="input"] > div,
    div[data-baseweb="input"] input,
    div[role="combobox"],
    div[role="combobox"] > div,
    div[role="combobox"] span,
    .stSelectbox div[data-baseweb="select"],
    .stSelectbox div[data-baseweb="select"] > div,
    .stSelectbox div[data-baseweb="select"] span,
    div[data-testid="stTextInput"] input {{
        background-color: {css_widget_bg} !important;
        background: {css_widget_bg} !important;
        color: {css_text_primary} !important;
        -webkit-text-fill-color: {css_text_primary} !important;
        fill: {css_text_primary} !important;
    }}
    
    div[data-baseweb="select"],
    div[data-baseweb="input"] > div,
    div[role="combobox"],
    div[data-testid="stTextInput"] input {{
        border: 1px solid {css_widget_border} !important;
        border-radius: 8px !important;
    }}

    /* Popover Dropdown Menu Items */
    div[data-baseweb="popover"] ul,
    div[data-baseweb="menu"],
    div[role="listbox"] {{
        background-color: {css_popover_bg} !important;
        border: 1px solid {css_widget_border} !important;
    }}

    div[data-baseweb="popover"] li,
    div[data-baseweb="menu"] li,
    div[role="option"] {{
        color: {css_text_primary} !important;
        background-color: {css_popover_bg} !important;
    }}

    div[data-baseweb="popover"] li:hover,
    div[data-baseweb="menu"] li:hover,
    div[role="option"]:hover {{
        background-color: {css_accent_purple} !important;
        color: #FFFFFF !important;
    }}

    /* Streamlit Button Styling - Green & Purple Theme */
    div.stButton > button,
    div.stButton > button[kind="secondary"] {{
        background: {css_btn_sec_bg} !important;
        color: {css_text_primary} !important;
        border: 1px solid {css_widget_border} !important;
        border-radius: 10px !important;
        padding: 8px 16px !important;
        font-weight: 600 !important;
        font-size: 15px !important;
        min-height: 56px !important;
        transition: all 0.3s cubic-bezier(0.16, 1, 0.3, 1) !important;
    }}
    
    div.stButton > button[kind="primary"] {{
        background: linear-gradient(135deg, {css_accent_green} 0%, {css_accent_purple} 100%) !important;
        color: #FFFFFF !important;
        border: 1px solid #10B981 !important;
        box-shadow: 0 8px 20px -4px rgba(16, 185, 129, 0.4), 0 0 14px 2px rgba(139, 92, 246, 0.3) !important;
        font-weight: 700 !important;
    }}
    
    div.stButton > button:hover {{
        transform: translateY(-2px) scale(1.01) !important;
        border-color: {css_accent_green} !important;
    }}
    
    /* Custom Card Style - Apple Liquid Glass Style (Dark Version) */
    .glass-card {{
        background: {css_card_bg} !important;
        backdrop-filter: blur(12px);
        -webkit-backdrop-filter: blur(12px);
        padding: 16px 20px !important;
        border-radius: 12px !important;
        border: 1px solid {css_card_border} !important;
        box-shadow: 0 4px 30px rgba(0, 0, 0, 0.3) !important;
        margin-bottom: 15px;
        transition: all 0.4s cubic-bezier(0.16, 1, 0.3, 1);
        position: relative;
        overflow: hidden;
        color: {css_text_primary} !important;
    }}
    
    /* Liquid glass glow on hover */
    .glass-card:hover {{
        transform: translateY(-3px);
        box-shadow: 0 16px 24px -8px rgba(16, 185, 129, 0.2), 0 8px 16px -6px rgba(139, 92, 246, 0.1);
        border-color: {css_accent_green} !important;
        background: {css_card_bg} !important;
    }}
    
    /* Sweep shine effect */
    .glass-card::before {{
        content: '';
        position: absolute;
        top: 0;
        left: -100%;
        width: 100%;
        height: 100%;
        background: linear-gradient(
            90deg,
            transparent,
            rgba(255, 255, 255, 0.05),
            transparent
        );
        transition: 0.6s ease-in-out;
    }}
    
    .glass-card:hover::before {{
        left: 100%;
    }}
    
    /* Streamlit Button Styling - Liquid Glass Apple Style */
    div.stButton > button,
    div.stButton > button[kind="primary"],
    div.stButton > button[kind="secondary"] {{
        background: {css_btn_sec_bg} !important;
        color: {css_text_primary} !important;
        border: 1px solid {css_widget_border} !important;
        border-radius: 10px !important;
        padding: 8px 16px !important;
        font-weight: 600 !important;
        font-size: 16px !important;
        min-height: 56px !important;
        display: inline-flex !important;
        align-items: center !important;
        justify-content: center !important;
        text-align: center !important;
        transition: all 0.3s cubic-bezier(0.16, 1, 0.3, 1) !important;
        position: relative !important;
        overflow: hidden !important;
        box-shadow: 0 4px 12px -2px rgba(16, 185, 129, 0.15) !important;
    }}
    
    /* Active Menu Button (Primary) */
    div.stButton > button[kind="primary"] {{
        background: linear-gradient(135deg, {css_accent_green} 0%, {css_accent_purple} 100%) !important;
        color: #FFFFFF !important;
        border: 1px solid #10B981 !important;
        box-shadow: 0 8px 24px -4px rgba(16, 185, 129, 0.4), 0 0 16px 4px rgba(139, 92, 246, 0.2) !important;
        font-weight: 700 !important;
    }}
    
    /* Hover Effects */
    div.stButton > button:hover {{
        transform: translateY(-3px) scale(1.02) !important;
        box-shadow: 0 12px 28px -4px rgba(16, 185, 129, 0.4) !important;
        border-color: {css_accent_green} !important;
    }}
    
    div.stButton > button[kind="primary"]:hover {{
        box-shadow: 0 12px 28px -4px rgba(16, 185, 129, 0.5), 0 0 20px 6px rgba(139, 92, 246, 0.4) !important;
        border-color: #10B981 !important;
    }}
    
    /* Click Effect */
    div.stButton > button:active {{
        transform: translateY(1px) scale(0.98) !important;
        filter: brightness(0.9) !important;
    }}
    
    /* Sweep shine effect on hover */
    div.stButton > button::before {{
        content: '' !important;
        position: absolute !important;
        top: 0 !important;
        left: -100% !important;
        width: 100% !important;
        height: 100% !important;
        background: linear-gradient(
            90deg,
            transparent,
            rgba(255, 255, 255, 0.15),
            transparent
        ) !important;
        transition: 0.5s !important;
    }}
    
    div.stButton > button:hover::before {{
        left: 100% !important;
    }}
    
    .accent-border-green {{
        border-left: 5px solid #10B981 !important;
    }}
    
    .accent-border-orange {{
        border-left: 5px solid #EF4444 !important;
    }}
    
    .accent-border-blue {{
        border-left: 5px solid #3B82F6 !important;
    }}
    
    /* Checklist badge styles */
    .badge-success {{
        background-color: rgba(16, 185, 129, 0.2) !important;
        color: #34D399 !important;
        padding: 3px 8px;
        border-radius: 6px;
        font-weight: bold;
        font-size: 0.85em;
        border: 1px solid rgba(16, 185, 129, 0.3) !important;
    }}
    
    .badge-warning {{
        background-color: rgba(245, 158, 11, 0.2) !important;
        color: #FBBF24 !important;
        padding: 3px 8px;
        border-radius: 6px;
        font-weight: bold;
        font-size: 0.85em;
        border: 1px solid rgba(245, 158, 11, 0.3) !important;
    }}
    
    .badge-danger {{
        background-color: rgba(239, 68, 68, 0.2) !important;
        color: #F87171 !important;
        padding: 3px 8px;
        border-radius: 6px;
        font-weight: bold;
        font-size: 0.85em;
        border: 1px solid rgba(239, 68, 68, 0.3) !important;
    }}
    
    .badge-info {{
        background-color: rgba(59, 130, 246, 0.2) !important;
        color: #60A5FA !important;
        padding: 3px 8px;
        border-radius: 6px;
        font-weight: bold;
        font-size: 0.85em;
        border: 1px solid rgba(59, 130, 246, 0.3) !important;
        margin-left: 0px;
        white-space: nowrap !important;
        display: inline-block !important;
    }}
    
    /* Log block style */
    .log-box {{
        background-color: #0F172A;
        color: #38BDF8;
        padding: 10px;
        border-radius: 6px;
        font-family: 'Consolas', monospace;
        font-size: 0.9em;
        max-height: 200px;
        overflow-y: auto;
        border: 1px solid rgba(255, 255, 255, 0.05);
    }}
    
    /* Compact spacing for checkboxes and inputs */
    div[data-testid="stCheckbox"] {{
        margin-bottom: -15px !important;
    }}
    
    div[data-testid="stTextInput"] input, 
    div[data-testid="stSelectbox"] select, 
    div[data-testid="stTextArea"] textarea {{
        background-color: {css_widget_bg} !important;
        color: {css_text_primary} !important;
        border: 1px solid {css_widget_border} !important;
        border-radius: 8px !important;
    }}
    
    div[data-testid="stTextInput"] > div, 
    div[data-testid="stSelectbox"] > div {{
        min-height: 32px !important;
    }}
    
    /* Lock sidebar open, hide collapse control trigger */
    [data-testid="collapsedControl"],
    button[data-testid="collapsedControl"],
    div[data-testid="collapsedControl"] button,
    [data-testid="stSidebarCollapseButton"],
    button[aria-label="Collapse sidebar"],
    button[aria-label="Expand sidebar"] {{
        display: none !important;
    }}
    
    /* Flexbox layout for sidebar content to push footer to the very bottom edge */
    section[data-testid="stSidebar"],
    section[data-testid="stSidebar"] > div:first-child,
    div[data-testid="stSidebarContent"],
    [data-testid="stSidebarUserContent"] {{
        display: flex !important;
        flex-direction: column !important;
        flex: 1 1 auto !important;
        height: 100% !important;
        min-height: calc(100vh - 20px) !important;
        padding-top: 0px !important;
    }}

    .sidebar-footer-pinned {{
        margin-top: auto !important;
        padding-top: 25px !important;
        padding-bottom: 15px !important;
    }}
</style>
""", unsafe_allow_html=True)

# Helper Functions
def get_parsed_datetime(date_str):
    """
    Parses RFC 822 / MIME date strings into a offset-naive datetime object for sorting.
    """
    if not date_str:
        return datetime.datetime.min
    try:
        from email.utils import parsedate_to_datetime
        dt = parsedate_to_datetime(date_str)
        if dt:
            return dt.replace(tzinfo=None)
    except Exception:
        pass
    try:
        return datetime.datetime.strptime(str(date_str)[:10], "%Y-%m-%d")
    except Exception:
        return datetime.datetime.min

def format_display_date(date_str):
    """
    Formats raw email date string into clean DD/MM/YYYY HH:MM format for display.
    """
    if not date_str:
        return "-"
    try:
        from email.utils import parsedate_to_datetime
        dt = parsedate_to_datetime(date_str)
        if dt:
            return dt.strftime("%d/%m/%Y %H:%M")
    except Exception:
        pass
    return str(date_str)[:16]
def clean_base_subject(subj):
    """
    Strips leading Re:, Fwd:, ตอบกลับ:, ส่งต่อ:, FW:, FWD:, etc. to extract the clean base thread subject.
    """
    if not subj:
        return ""
    prev = None
    curr = subj.strip()
    while prev != curr:
        prev = curr
        curr = re.sub(r'^(?:re|fwd?|fw|ตอบกลับ|ส่งต่อ|ส่งแก้ไข)\s*:\s*', '', curr, flags=re.IGNORECASE).strip()
    return curr

def extract_hospital_name(subject, body, filenames=None):
    """
    Attempts to extract a clean hospital or laboratory name from the subject line first,
    then body, and finally attachment filenames.
    Cleanly strips trailing province/district markers (e.g., จ., จังหวัด, อำเภอ).
    """
    def clean_name(raw_name):
        if not raw_name:
            return ""
        # Remove line breaks and normalize spaces
        raw_name = raw_name.replace("\r", " ").replace("\n", " ").replace("\t", " ")
        raw_name = " ".join(raw_name.split())
        # Strip trailing province or district suffixes
        raw_name = re.sub(r'\s*(จังหวัด|จ\.|จ\b|อำเภอ|อ\.).*$', '', raw_name).strip()
        # If it ends with a stray single 'จ' or dot
        if raw_name.endswith(" จ"):
            raw_name = raw_name[:-2].strip()
        elif raw_name.endswith("จ") and not raw_name.endswith("เวช"):
            raw_name = raw_name[:-1].strip()
        return raw_name

    # 1. Search inside the subject line
    if subject:
        m = re.search(r'(โรงพยาบาล\s*[ก-๙a-zA-Z0-9_]+(?:\s+[ก-๙a-zA-Z0-9_]+)*)', subject)
        if m:
            name = clean_name(m.group(1))
            if len(name) >= 8:
                return name
                
        m = re.search(r'(รพ\.\s*[ก-๙a-zA-Z0-9_]+|รพ\s+[ก-๙a-zA-Z0-9_]+)', subject)
        if m:
            name = clean_name(m.group(1))
            if len(name) >= 3:
                return name

        m = re.search(r'([ก-๙a-zA-Z0-9_\s]+สหคลินิก|[ก-๙a-zA-Z0-9_\s]+คลินิก)', subject)
        if m:
            name = clean_name(m.group(1))
            if len(name) >= 3:
                return name

    # 2. Fallback to search inside the email body
    if body:
        m = re.search(r'(โรงพยาบาล\s*[ก-๙a-zA-Z0-9_]+(?:\s+[ก-๙a-zA-Z0-9_]+)*)', body)
        if m:
            name = clean_name(m.group(1))
            if len(name) >= 8:
                return name
                
        m = re.search(r'(รพ\.\s*[ก-๙a-zA-Z0-9_]+|รพ\s+[ก-๙a-zA-Z0-9_]+)', body)
        if m:
            name = clean_name(m.group(1))
            if len(name) >= 3:
                return name

        m = re.search(r'([ก-๙a-zA-Z0-9_\s]+สหคลินิก|[ก-๙a-zA-Z0-9_\s]+คลินิก)', body)
        if m:
            name = clean_name(m.group(1))
            if len(name) >= 3:
                return name

    # 3. Fallback to search inside attachment filenames
    if filenames:
        for fn in filenames:
            m = re.search(r'(โรงพยาบาล\s*[ก-๙a-zA-Z0-9_]+(?:\s+[ก-๙a-zA-Z0-9_]+)*)', fn)
            if m:
                name = clean_name(m.group(1))
                if len(name) >= 8:
                    return name
            m = re.search(r'(รพ\.\s*[ก-๙a-zA-Z0-9_]+|รพ\s+[ก-๙a-zA-Z0-9_]+)', fn)
            if m:
                name = clean_name(m.group(1))
                if len(name) >= 3:
                    return name
            fn_name = os.path.splitext(fn)[0]
            m2 = re.search(r'(?:V\.\d|Checklist|user)\s*([ก-๙a-zA-Z0-9_]+)', fn_name)
            if m2:
                word = clean_name(m2.group(1))
                if len(word) >= 3 and not word.isdigit():
                    if "โรงพยาบาล" not in word and "รพ" not in word:
                        return f"รพ.{word}"
                    return word

    return "ไม่พบชื่อโรงพยาบาล"

def normalize_hospital_key(name):
    """
    Normalizes hospital names into a clean key for strict cross-round folder matching and merging.
    Example: '2026-08-05_โรงพยาบาลนาแก' -> 'นาแก', 'รพ.นาแก' -> 'นาแก'
    """
    if not name:
        return ""
    n = str(name).strip()
    n = re.sub(r'^\d{4}-\d{2}-\d{2}_', '', n)
    n = re.sub(r'^(โรงพยาบาล|รพ\.|รพ\s*)', '', n).strip()
    n = re.sub(r'[\s\.\-_\(\)]', '', n).lower()
    return n

def get_hospital_folder_path(hospital_name, date_str, is_complete=False):
    """
    Resolves hospital folder path under:
      '[workspace_dir]/2. รพ ที่ verified แล้ว' (if complete)
      '[workspace_dir]/1. นำเข้าใหม่ยังไม่ได้เช็ค' (if incomplete)
    Merges/moves existing folders dynamically if completeness status changes or if hospital sent multiple rounds.
    """
    workspace = st.session_state.get("workspace_dir", os.path.abspath(os.path.dirname(__file__)))
    dir_complete = os.path.join(workspace, "2. รพ ที่ verified แล้ว")
    dir_incomplete = os.path.join(workspace, "1. นำเข้าใหม่ยังไม่ได้เช็ค")
    os.makedirs(dir_complete, exist_ok=True)
    os.makedirs(dir_incomplete, exist_ok=True)

    hospital_clean = re.sub(r'[\\/*?:"<>|]', "_", hospital_name or "หน่วยงานนิรนาม").strip()
    target_parent = dir_complete if is_complete else dir_incomplete

    found_path = None
    found_parent = None
    found_foldername = None

    target_key = normalize_hospital_key(hospital_clean)

    for parent_dir in [dir_complete, dir_incomplete]:
        if os.path.exists(parent_dir):
            for f in os.listdir(parent_dir):
                f_key = normalize_hospital_key(f)
                if target_key and f_key and (target_key == f_key or target_key in f_key or f_key in target_key):
                    found_path = os.path.join(parent_dir, f)
                    found_parent = parent_dir
                    found_foldername = f
                    break
            if found_path:
                break

    if found_path:
        if found_parent != target_parent:
            new_path = os.path.join(target_parent, found_foldername)
            try:
                if os.path.exists(new_path):
                    for item in os.listdir(found_path):
                        s = os.path.join(found_path, item)
                        d = os.path.join(new_path, item)
                        if os.path.isfile(s):
                            shutil.copy2(s, d)
                    shutil.rmtree(found_path, ignore_errors=True)
                    return new_path
                else:
                    shutil.move(found_path, new_path)
                    return new_path
            except Exception:
                return found_path
        return found_path
    else:
        new_foldername = f"{date_str}_{hospital_clean}"
        new_path = os.path.join(target_parent, new_foldername)
        os.makedirs(new_path, exist_ok=True)
        return new_path

def sort_hospital_records(items, sort_option):
    """
    Sorts a list of hospital records based on user choice.
    """
    def get_sort_key(rec):
        if sort_option == "📅 วันที่เสนอตรวจประเมิน":
            val = str(rec.get("appointment", "") or "").strip()
            return (val == "", val)
        elif sort_option == "⏳ วันครบกำหนดใบอนุญาตหมดอายุ":
            val = str(rec.get("expiry_date", "") or "").strip()
            return (val == "", val)
        elif sort_option == "🏥 ชื่อหน่วยงาน/โรงพยาบาล (A-Z)":
            return str(rec.get("hospital_name", "") or "").lower()
        else: # "🕒 ลำดับการบันทึก (ล่าสุดขึ้นก่อน)"
            return str(rec.get("last_updated", "") or "")
            
    reverse_sort = (sort_option == "🕒 ลำดับการบันทึก (ล่าสุดขึ้นก่อน)")
    return sorted(items, key=get_sort_key, reverse=reverse_sort)

def browse_directory():
    """
    Opens a native Tkinter file dialog to browse local directories.
    """
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        dir_path = filedialog.askdirectory(
            title="เลือกโฟลเดอร์เก็บข้อมูลโครงการ"
        )
        root.destroy()
        return dir_path
    except Exception:
        return ""

def log_usage(operator_name, action):
    """
    Logs operator actions in usage_logs.csv for audit trail/statistics.
    """
    file_exists = os.path.exists(USAGE_LOG_FILE)
    try:
        with open(USAGE_LOG_FILE, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if not file_exists:
                writer.writerow(["Timestamp", "Operator Name", "Action"])
            writer.writerow([datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), operator_name, action])
    except Exception as e:
        print(f"Error logging usage: {e}")

def zip_folder(folder_path):
    """
    Zips a folder dynamically in-memory and returns a BytesIO object for downloading.
    """
    zip_buffer = io.BytesIO()
    if os.path.exists(folder_path):
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for root, dirs, files in os.walk(folder_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, folder_path)
                    zip_file.write(file_path, arcname)
    zip_buffer.seek(0)
    return zip_buffer

def copy_files_with_conflict_check(src_dir, dest_dir, conflict_behavior):
    """
    Copies files from temp directory to destination with conflict behavior.
    """
    if not os.path.exists(src_dir):
        return []
        
    os.makedirs(dest_dir, exist_ok=True)
    logs = []
    
    for filename in os.listdir(src_dir):
        src_file = os.path.join(src_dir, filename)
        if not os.path.isfile(src_file):
            continue
            
        dest_file = os.path.join(dest_dir, filename)
        
        if os.path.exists(dest_file):
            if "เขียนทับ" in conflict_behavior:
                shutil.copy2(src_file, dest_file)
                logs.append(f"🔄 เขียนทับไฟล์แนบเดิม: {filename}")
            elif "บันทึกคู่กัน" in conflict_behavior:
                name, ext = os.path.splitext(filename)
                counter = 1
                new_dest = os.path.join(dest_dir, f"{name}_ใหม่{counter}{ext}")
                while os.path.exists(new_dest):
                    counter += 1
                    new_dest = os.path.join(dest_dir, f"{name}_ใหม่{counter}{ext}")
                shutil.copy2(src_file, new_dest)
                logs.append(f"📄 บันทึกคู่กันเป็น: {os.path.basename(new_dest)}")
            else: # ข้าม
                logs.append(f"⏭️ ข้ามไฟล์แนบซ้ำ: {filename}")
        else:
            shutil.copy2(src_file, dest_file)
            logs.append(f"📥 บันทึกไฟล์แนบใหม่: {filename}")
            
    return logs

# Load Configuration Data
def load_config():
    config_data = {}
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                config_data = json.load(f)
        except Exception:
            pass
            
    # Default Config Values
    config_data.setdefault("email", "")
    config_data.setdefault("password", "")
    config_data.setdefault("imap_server", "imap.gmail.com")
    config_data.setdefault("api_key", "")
    
    # Default Workspace directory
    default_workspace = os.path.abspath(os.path.dirname(__file__))
    curr_ws = config_data.get("workspace_dir", "")
    if not curr_ws or not os.path.exists(curr_ws):
        config_data["workspace_dir"] = default_workspace

    # Derive Excel path
    config_data["excel_path"] = os.path.join(config_data["workspace_dir"], "ตารางตรวจ LA สิงหาคม 69.xlsx")
    config_data.setdefault("download_attachments", True)
    return config_data

def save_config(config_data):
    try:
        os.makedirs(os.path.dirname(os.path.abspath(CONFIG_FILE)), exist_ok=True)
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(config_data, f, ensure_ascii=False, indent=4)
        return True
    except Exception:
        return False

# Load Registry Data
def load_hospital_registry():
    if os.path.exists(HOSPITAL_REGISTRY_FILE):
        try:
            with open(HOSPITAL_REGISTRY_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def save_hospital_registry(registry):
    try:
        with open(HOSPITAL_REGISTRY_FILE, "w", encoding="utf-8") as f:
            json.dump(registry, f, ensure_ascii=False, indent=4)
        return True
    except Exception:
        return False

# Load Assessors
def load_assessors():
    excel_assessor_path = "หน่วยงานที่ขอตรวจ.xlsx"
    if os.path.exists(excel_assessor_path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_assessor_path, data_only=True)
            sheet_name = wb.sheetnames[1] if len(wb.sheetnames) > 1 else wb.sheetnames[0]
            sh = wb[sheet_name]
            
            header_row = next(sh.iter_rows(max_row=1, values_only=True))
            name_col_idx = 1
            for idx, val in enumerate(header_row):
                if val and "ชื่อ" in str(val) and "สกุล" in str(val):
                    name_col_idx = idx
                    break
                    
            names = []
            for row in sh.iter_rows(min_row=2, values_only=True):
                if len(row) > name_col_idx and row[name_col_idx]:
                    clean_name = " ".join(str(row[name_col_idx]).split())
                    if clean_name:
                        names.append(clean_name)
            if names:
                return names
        except Exception:
            pass

    if os.path.exists(ASSESSORS_FILE):
        try:
            with open(ASSESSORS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
            
    # Fallback default list
    return [
        "ดร. นงลักษณ์ สมบูรณ์", "นาย สาธิต รักษ์ดี", "นาง วิรุฬห์ เกียรติกร",
        "นาย ประสงค์ มั่นคง", "นางสาว อรทัย เลิศปัญญา", "นาง สมจิต รักเรียน",
        "นาย สมชาย ใจดี", "นาง นภา เลิศเกียรติ", "นาย วีระพล รักษ์ดี",
        "นางสาว สิริมา วงศ์ทอง", "ทนพ. พิเชษฐ์ เก่งการค้า", "ทนพญ. รัชนี อิตตกี",
        "ดร.ทนพ. สมเกียรติ ยอดรัก", "ทนพ. อานนท์ วงศ์สว่าง", "ทนพญ. วรรณภา ศรีสุข",
        "ทนพ. ปริญญา มั่นคง", "ทนพญ. นงนุช เจริญดี", "ทนพ. มานพ รักษาการ",
        "ทนพญ. จิตราภรณ์ แสนดี", "ทนพญ. สุรีย์พร แดงทองดี", "ทนพ. สุรศักดิ์ วิชัยดิษฐ์",
        "ทนพญ. พรทิพย์ สุวรรณรัตน์", "ทนพ. เกรียงไกร ชัยชนะ", "ทนพญ. ณิชชา วงศ์สุวัฒน์",
        "ทนพ. ธีรเดช เจริญพงษ์", "ทนพญ. นิภาวรรณ สมมิตร", "ทนพ. ชาญชัย วรโชติ",
        "ทนพญ. จารุวรรณ สินธุวงศ์", "ทนพ. ประวิทย์ อรุณรุ่ง", "ทนพญ. วาสนา รุ่งเรือง"
    ]

# Initialize Session State
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
if "user_info" not in st.session_state:
    st.session_state.user_info = None

# Gating page access based on auth status
if not st.session_state.logged_in:
    render_login_page()
    st.stop()

if "operator_name" not in st.session_state:
    st.session_state.operator_name = st.session_state.user_info.get("fullname", "ผู้ใช้งานทั่วไป") if st.session_state.user_info else "ผู้ใช้งานทั่วไป"
if "config" not in st.session_state:
    st.session_state.config = load_config()
if "registry" not in st.session_state:
    st.session_state.registry = load_hospital_registry()
if "assessors" not in st.session_state:
    st.session_state.assessors = load_assessors()
if "fetched_emails" not in st.session_state:
    st.session_state.fetched_emails = []
if "active_menu" not in st.session_state:
    st.session_state.active_menu = 1
if "workspace_dir" not in st.session_state:
    st.session_state.workspace_dir = st.session_state.config.get("workspace_dir", "")
if "conflict_behavior" not in st.session_state:
    st.session_state.conflict_behavior = "เขียนทับด้วยไฟล์ใหม่ (Overwrite)"

cfg = st.session_state.config

# ----------------- APP SIDEBAR (CONFIGURATIONS & GENERAL SETTINGS) -----------------
with st.sidebar:
    user = st.session_state.user_info
    disp_name = user.get("display_name", "ผู้ใช้งาน") if user else "ผู้ใช้งาน"
    full_name = user.get("fullname", "") if user else ""
    user_role = user.get("role", "") if user else ""
    
    st.markdown(f"""
    <div style='padding: 14px 10px; background: rgba(255,255,255,0.03); border: 1px solid rgba(16, 185, 129, 0.2); border-radius: 12px; box-shadow: 0 4px 20px 0 rgba(0,0,0,0.15); text-align: center;'>
        <div style='background: linear-gradient(135deg, #10B981 0%, #8B5CF6 100%); -webkit-background-clip: text; -webkit-text-fill-color: transparent; margin: 0; font-size: 22px !important; font-weight: 900; line-height: 1.3;'>Laboratory Accreditation Management System (LAMS)</div>
    </div>
    <div style='color: {css_text_secondary}; font-size: 8.5px !important; font-weight: 300 !important; font-style: italic; text-align: center; margin-top: 6px; margin-bottom: 14px; line-height: 1.3;'>Integrated Quality, Evidence & Assessment Management Platform</div>
    
    <div style='padding: 12px; background: rgba(16, 185, 129, 0.08); border: 1px solid rgba(16, 185, 129, 0.25); border-radius: 8px; text-align: center; margin-bottom: 15px;'>
        <div style='color: #10B981; font-size: 16px !important; font-weight: 700;'>สวัสดี! {disp_name}</div>
        <div style='color: #059669; font-size: 14px !important; font-weight: 600; margin-top: 4px;'>{full_name}</div>
        <div style='color: #8B5CF6; font-size: 10px !important; font-weight: 400; margin-top: 6px; line-height: 1.4;'>{user_role}</div>
    </div>
    """, unsafe_allow_html=True)
    
    with st.expander("⚙️ ตั้งค่าระบบ (System Settings)"):
        # Theme Mode Selector
        st.markdown("#### 🎨 ธีมแสดงผล (Theme Mode)")
        theme_choice = st.selectbox(
            "เลือกโหมดสีหน้าจอ:",
            ["🌙 Dark Mode", "☀️ Bright / Light Mode"],
            index=0 if st.session_state.theme_mode == "🌙 Dark Mode" else 1,
            key="theme_mode_selectbox"
        )
        if theme_choice != st.session_state.theme_mode:
            st.session_state.theme_mode = theme_choice
            st.rerun()
            
        with st.expander("🔑 เปลี่ยนรหัสผ่าน (Change Password)", expanded=False):
            curr_pwd = st.text_input("รหัสผ่านเดิม:", type="password", key="chg_curr_pwd")
            new_pwd = st.text_input("รหัสผ่านใหม่:", type="password", key="chg_new_pwd")
            confirm_pwd = st.text_input("ยืนยันรหัสผ่านใหม่:", type="password", key="chg_conf_pwd")
            
            if st.button("💾 ยืนยันเปลี่ยนรหัสผ่าน", use_container_width=True, key="btn_confirm_chg_pwd"):
                if not curr_pwd or not new_pwd or not confirm_pwd:
                    st.error("กรุณากรอกข้อมูลให้ครบทุกช่อง!")
                elif new_pwd != confirm_pwd:
                    st.error("รหัสผ่านใหม่กับยืนยันรหัสผ่านไม่ตรงกัน!")
                else:
                    users = load_users()
                    u_key = user.get("username", "").strip().lower() if user else ""
                    u_obj = users.get(u_key)
                    if u_obj:
                        check_hash, _ = hash_password(curr_pwd, u_obj["salt"])
                        if check_hash == u_obj["password_hash"]:
                            new_hash, new_salt = hash_password(new_pwd)
                            u_obj["password_hash"] = new_hash
                            u_obj["salt"] = new_salt
                            users[u_key] = u_obj
                            if save_users(users):
                                st.session_state.user_info = u_obj
                                st.success("🎉 เปลี่ยนรหัสผ่านสำเร็จ!")
                                log_usage(st.session_state.operator_name, f"เปลี่ยนรหัสผ่านของบัญชี {u_key}")
                                time.sleep(0.5)
                                st.rerun()
                            else:
                                st.error("เกิดข้อผิดพลาดในการบันทึกข้อมูลลง users.json")
                        else:
                            st.error("รหัสผ่านเดิมไม่ถูกต้อง!")
                    else:
                        st.error("ไม่พบบัญชีผู้ใช้งานนี้ในระบบ!")
                    
        st.markdown("---")
        st.markdown("#### ⚙️ การตั้งค่าระบบดึงข้อมูล")
        email = st.text_input("อีเมลดึงจดหมาย (Gmail):", value=cfg.get("email", ""), key="cfg_email")
        password = st.text_input("รหัสผ่านแอป (App Password):", value=cfg.get("password", ""), type="password", help="⚠️ ไม่ใช่รหัสผ่านปกติของเมล! ได้มาจาก:\n1. เปิดระบบยืนยัน 2 ขั้นตอน (2-Step Verification) ในบัญชี Google\n2. เข้าหน้าเว็บ myaccount.google.com/security\n3. ค้นหาคำว่า 'App Passwords' (รหัสผ่านสำหรับแอป)\n4. กดสร้างรหัสผ่านใหม่และคัดลอกรหัสผ่าน 16 ตัวมาระบุในช่องนี้", key="cfg_pwd")
        api_key = st.text_input("Gemini API Key:", value=cfg.get("api_key", ""), type="password", help="ใช้ในการส่งเนื้อความอีเมลและ PDF ไปประมวลผลดึงโครงสร้าง JSON", key="cfg_apikey")
        
        # Stacked Folder Selector
        workspace_dir_input = st.text_input("โฟลเดอร์เก็บข้อมูลโครงการ:", value=st.session_state.workspace_dir, key="cfg_wsdir")
        if st.button("📂 คลิกเพื่อเลือกโฟลเดอร์บนเครื่อง...", use_container_width=True, key="btn_cfg_browse"):
            selected_dir = browse_directory()
            if selected_dir:
                st.session_state.workspace_dir = selected_dir
                st.rerun()
                    
        if st.button("💾 บันทึกการตั้งค่าลงเครื่อง", use_container_width=True, type="primary", key="btn_cfg_save"):
            cfg["email"] = email.strip()
            cfg["password"] = password.strip()
            cfg["api_key"] = api_key.strip()
            cfg["workspace_dir"] = workspace_dir_input.strip()
            cfg["excel_path"] = os.path.join(cfg["workspace_dir"], "ตารางตรวจ LA สิงหาคม 69.xlsx")
            cfg["download_attachments"] = True
            
            if save_config(cfg):
                st.session_state.config = cfg
                st.session_state.workspace_dir = workspace_dir_input.strip()
                st.success("บันทึกการตั้งค่าสำเร็จ!")
                log_usage(st.session_state.operator_name, "บันทึกการตั้งค่า config.json ใหม่")
                time.sleep(0.5)
                st.rerun()
            else:
                st.error("เกิดข้อผิดพลาดในการเขียนไฟล์ config.json")
                
        st.markdown("---")
        if st.button("🚪 ออกจากระบบ (Logout)", use_container_width=True, key="btn_logout_inside"):
            st.session_state.logged_in = False
            st.session_state.user_info = None
            st.session_state.operator_name = "ผู้ใช้งานทั่วไป"
            st.rerun()

    st.markdown(
        f"<div class='sidebar-footer-pinned'>"
        f"<div style='text-align: center; color: {css_text_secondary}; font-size: 5px !important; font-family: \"TH Sarabun New\", \"TH SarabunPSK\", \"Sarabun\", sans-serif !important; font-style: italic !important; line-height: 1.3; background: rgba(0,0,0,0.02); padding: 3px 2px; border-radius: 5px; border: 1px solid rgba(128,128,128,0.12);'>"
        f"<span style='color: {css_text_secondary}; font-size: 5px !important; font-style: italic; font-family: \"TH Sarabun New\", \"TH SarabunPSK\", \"Sarabun\", sans-serif !important;'>พัฒนาโดย</span><br>"
        f"<span style='color: {css_text_primary}; font-size: 5px !important; font-weight: 600; font-style: italic; white-space: nowrap !important; font-family: \"TH Sarabun New\", \"TH SarabunPSK\", \"Sarabun\", sans-serif !important;'>ทนพ.ปองพล ฤกษ์เนาวรัตน์</span><br>"
        f"<span style='color: {css_text_secondary}; font-size: 5px !important; font-style: italic; font-family: \"TH Sarabun New\", \"TH SarabunPSK\", \"Sarabun\", sans-serif !important;'>นักเทคนิคการแพทย์ ชำนาญการ</span>"
        f"</div>"
        f"</div>",
        unsafe_allow_html=True
    )

# ----------------- CUSTOM NAVIGATION MENU BAR -----------------
# 2 columns side-by-side with liquid glass hover effects, spanning full screen width
col_m1, col_m2 = st.columns(2)

m1_clicked = col_m1.button(
    "📥 1. สำรวจและนำเข้าอีเมล", 
    key="m1_btn", 
    use_container_width=True, 
    type="primary" if st.session_state.active_menu == 1 else "secondary"
)
m2_clicked = col_m2.button(
    "🔍 2. วิเคราะห์และตรวจสอบความถูกต้องของเอกสาร", 
    key="m2_btn", 
    use_container_width=True, 
    type="primary" if st.session_state.active_menu == 2 else "secondary"
)

if m1_clicked:
    st.session_state.active_menu = 1
    st.rerun()
if m2_clicked:
    st.session_state.active_menu = 2
    st.rerun()

st.markdown("---")

# ----------------- MENU CONTENT RENDERING -----------------

# ================= MENU 1: FETCH & QUICK IMPORT =================
if st.session_state.active_menu == 1:
    if "show_success_modal" in st.session_state and st.session_state.show_success_modal:
        count, total_cnt = st.session_state.show_success_modal
        st.session_state.show_success_modal = None
        show_success_modal(count, total_cnt)
        
    st.subheader("📥 เมนูที่ 1: สำรวจอีเมลและนำเข้าเอกสารจากอีเมล")
    
    col_sel, col_btn = st.columns([2.5, 2.5])
    with col_sel:
        range_val = st.selectbox(
            "ช่วงเวลาในการดึงอีเมล:", 
            [
                "5 อีเมลล่าสุด", 
                "10 อีเมลล่าสุด", 
                "1 วันย้อนหลัง", 
                "3 วันย้อนหลัง", 
                "1 สัปดาห์ย้อนหลัง", 
                "1 เดือนย้อนหลัง", 
                "1 ปีย้อนหลัง", 
                "เฉพาะที่ยังไม่ได้อ่าน (ทั้งหมด)"
            ]
        )
    with col_btn:
        st.write("<div style='height: 28px;'></div>", unsafe_allow_html=True)
        if st.button("📥 ดึงและอ่านอีเมลล่าสุดจาก Gmail", type="secondary", use_container_width=True):
            if not cfg.get("email") or not cfg.get("password"):
                st.error("กรุณาระบุอีเมลและรหัสผ่านแอป Gmail ในแทบตั้งค่าด้านซ้ายก่อน!")
            else:
                days_limit = 7
                max_emails = None
                
                if range_val == "5 อีเมลล่าสุด":
                    days_limit = 365
                    max_emails = 5
                elif range_val == "10 อีเมลล่าสุด":
                    days_limit = 365
                    max_emails = 10
                elif range_val == "1 วันย้อนหลัง":
                    days_limit = 1
                elif range_val == "3 วันย้อนหลัง":
                    days_limit = 3
                elif range_val == "1 สัปดาห์ย้อนหลัง":
                    days_limit = 7
                elif range_val == "1 เดือนย้อนหลัง":
                    days_limit = 30
                elif range_val == "1 ปีย้อนหลัง":
                    days_limit = 365
                elif "ยังไม่ได้อ่าน" in range_val:
                    days_limit = 0
                    
                with st.spinner("กำลังเชื่อมต่อเซิร์ฟเวอร์อีเมล..."):
                    try:
                        # Empty keywords matches all emails
                        emails = fetch_emails_by_range(
                            imap_server=cfg.get("imap_server", "imap.gmail.com"),
                            email_addr=cfg.get("email"),
                            password=cfg.get("password"),
                            subject_keywords=[],
                            days_limit=days_limit,
                            max_emails=max_emails
                        )
                        st.session_state.fetched_emails = emails
                        log_usage(st.session_state.operator_name, f"ค้นหาอีเมล ({range_val}) พบ {len(emails)} ฉบับ")
                        if emails:
                            st.success(f"ดึงจดหมายเสร็จสิ้น! พบจดหมาย {len(emails)} ฉบับ")
                        else:
                            st.warning("ไม่พบจดหมายใหม่ในระยะเวลาที่ระบุ")
                    except Exception as e:
                        st.error(f"เกิดข้อผิดพลาดในการดึงอีเมล: {e}")
        
    if st.session_state.fetched_emails:
        st.markdown("#### 📧 รายชื่ออีเมลที่พบ:")
        
        # Add Search Filter input
        search_query = st.text_input("🔍 กรองผลการค้นหา (หัวข้อจดหมาย / ชื่อผู้ส่ง):", placeholder="พิมพ์คำสำคัญเพื่อค้นหา เช่น la, ประเมิน, หรือชื่อโรงพยาบาล")
            
        # Apply filter
        if search_query.strip():
            filtered_emails = [
                e for e in st.session_state.fetched_emails
                if search_query.lower() in e['subject'].lower() or search_query.lower() in e['sender'].lower()
            ]
        else:
            filtered_emails = st.session_state.fetched_emails
            
        # Always sort emails by received date descending (newest on top)
        filtered_emails = sorted(filtered_emails, key=lambda e: get_parsed_datetime(e.get("date", "")), reverse=True)
            
        if not filtered_emails:
            st.warning("⚠️ ไม่พบอีเมลที่ตรงกับเงื่อนไขการกรอง")
            selected_ids = []
        else:
            selected_ids = []
            
            # --- Pre-processing for Thread Grouping & Hospital Name Propagation ---
            for e in st.session_state.fetched_emails:
                e["base_subject"] = clean_base_subject(e["subject"])
                fn_list = []
                if e.get("temp_dir") and os.path.exists(e["temp_dir"]):
                    try:
                        fn_list = os.listdir(e["temp_dir"])
                    except Exception:
                        pass
                e["parsed_h_name"] = extract_hospital_name(e["subject"], e.get("body", ""), fn_list)

            thread_h_map = {}
            for e in st.session_state.fetched_emails:
                bs = e["base_subject"]
                h = e["parsed_h_name"]
                if bs not in thread_h_map:
                    thread_h_map[bs] = "ไม่พบชื่อโรงพยาบาล"
                if h != "ไม่พบชื่อโรงพยาบาล" and thread_h_map[bs] == "ไม่พบชื่อโรงพยาบาล":
                    thread_h_map[bs] = h

            thread_files_map = {}
            thread_cloud_map = {}
            for e in st.session_state.fetched_emails:
                bs = e["base_subject"]
                h = thread_h_map.get(bs, e["parsed_h_name"])
                key = h if h != "ไม่พบชื่อโรงพยาบาล" else bs
                
                if key not in thread_files_map:
                    thread_files_map[key] = set()
                    thread_cloud_map[key] = False
                    
                if e.get("temp_dir") and os.path.exists(e["temp_dir"]):
                    try:
                        f_list = os.listdir(e["temp_dir"])
                        for f in f_list:
                            thread_files_map[key].add(f)
                            if f.endswith(".url"):
                                thread_cloud_map[key] = True
                    except Exception:
                        pass
            
            # Table Header - Clean select-all checkbox without squished text label
            h_col1, h_col2, h_col3, h_col4, h_col5, h_col6 = st.columns([0.6, 3.4, 2.4, 2.0, 2.0, 1.6])
            select_all = h_col1.checkbox("", value=True, key="select_all_header", label_visibility="collapsed")
            h_col2.markdown("**หัวข้อจดหมาย**")
            h_col3.markdown("**ไฟล์แนบ**")
            h_col4.markdown("**ผู้ส่ง**")
            h_col5.markdown("**ชื่อโรงพยาบาล**")
            h_col6.markdown("**วันที่ส่ง**")
            st.markdown("---")
            
            for email_item in filtered_emails:
                col1, col2, col3, col4, col5, col6 = st.columns([0.6, 3.4, 2.4, 2.0, 2.0, 1.6])
                
                bs = email_item.get("base_subject", clean_base_subject(email_item["subject"]))
                h_name = thread_h_map.get(bs, email_item.get("parsed_h_name", "ไม่พบชื่อโรงพยาบาล"))
                
                key = h_name if h_name != "ไม่พบชื่อโรงพยาบาล" else bs
                seen_thread_files = thread_files_map.get(key, set())
                has_cloud_link = thread_cloud_map.get(key, False)
                file_count = len(seen_thread_files)
                
                badge_style = "badge-success" if email_item['status'] == "อ่านแล้ว" else "badge-danger"
                cloud_badge = "<span class='badge-info' style='background: rgba(168, 85, 247, 0.2); color: #C084FC; border-color: rgba(168, 85, 247, 0.4); white-space: nowrap !important;'>🔗 Cloud Link</span>" if has_cloud_link else ""
                
                subject_text = f"{email_item['subject']} <span class='{badge_style}'>{email_item['status']}</span>"
                files_badge_text = f"<div style='display: flex; gap: 6px; align-items: center; white-space: nowrap !important;'><span class='badge-info'>📎 {file_count} ไฟล์</span>{cloud_badge}</div>"
                
                is_selected = col1.checkbox("", key=f"sel_mail_{email_item['id']}", value=select_all)
                col2.markdown(subject_text, unsafe_allow_html=True)
                col3.markdown(files_badge_text, unsafe_allow_html=True)
                col4.write(email_item['sender'][:30])
                col5.write(h_name)
                col6.write(format_display_date(email_item.get('date', '')))
                
                if is_selected:
                    selected_ids.append(email_item['id'])
                
        st.markdown("---")
        
        # Quick Import Action
        if st.button("📁 บันทึกไฟล์ ตามรายชื่อโรงพยาบาล", type="secondary", use_container_width=True):
            queue = [e for e in st.session_state.fetched_emails if e['id'] in selected_ids]
            if not queue:
                st.warning("กรุณาเลือกอีเมลอย่างน้อย 1 รายการก่อน!")
            elif not cfg.get("api_key"):
                st.error("กรุณากรอก Gemini API Key ในเมนูด้านซ้ายเพื่อส่งวิเคราะห์ AI!")
            else:
                progress_bar = st.progress(0)
                progress_text = st.empty()
                success_count = 0
                total = len(queue)
                
                for idx, mail in enumerate(queue):
                    percentage = int((idx) / total * 100)
                    progress_text.info(
                        f"⏳ **กำลังดาวน์โหลดและคัดแยก:** โหลดไปแล้ว {idx} จาก {total} โรงพยาบาล ({percentage}%)\n\n"
                        f"✉️ **จดหมายปัจจุบัน:** `{mail['subject'][:50]}...`"
                    )
                    progress_bar.progress(idx / total)
                    
                    try:
                        # 1. Resolve Hospital Name using fast regex extractor (Instant & 100% Free)
                        h_name = extract_hospital_name(mail["subject"], mail.get("body", "")).strip()
                        if not h_name or h_name == "ไม่พบชื่อโรงพยาบาล":
                            h_name = f"หน่วยงาน_Mail_{mail['id']}"
                            
                        email_date = parse_email_date_to_yyyy_mm_dd(mail["date"])
                        initial_comp = mail.get("completeness", {}).get("complete", False)
                        dest_dir = get_hospital_folder_path(h_name, email_date, initial_comp)
                        
                        # 2. ALWAYS copy attachments across ALL related emails for the same hospital (Guaranteed full thread capture!)
                        h_key = normalize_hospital_key(h_name)
                        related_emails = [
                            e for e in st.session_state.fetched_emails
                            if normalize_hospital_key(extract_hospital_name(e["subject"], e.get("body", ""))) == h_key
                        ]
                        if not related_emails:
                            related_emails = [mail]
                            
                        for rel_mail in related_emails:
                            if rel_mail.get("temp_dir") and os.path.exists(rel_mail["temp_dir"]):
                                copy_files_with_conflict_check(rel_mail["temp_dir"], dest_dir, st.session_state.conflict_behavior)
                                
                        # 2.1 Deep IMAP Thread Sweep: Query IMAP server to recover 100% of attachments across all forward/reply iterations
                        try:
                            swept_temp = sweep_thread_attachments_imap(mail, cfg)
                            if swept_temp and os.path.exists(swept_temp):
                                copy_files_with_conflict_check(swept_temp, dest_dir, st.session_state.conflict_behavior)
                        except Exception as sweep_err:
                            print(f"Thread Sweep Warning: {sweep_err}")
                            
                        # 3. Recalculate folder completeness locally
                        comp = check_folder_completeness(dest_dir)
                        dest_dir = get_hospital_folder_path(h_name, email_date, comp["complete"])
                        
                        # 4. Prepare data record
                        sender_email = ""
                        s_match = re.findall(r'[\w\.-]+@[\w\.-]+', mail["sender"])
                        if s_match:
                            sender_email = s_match[0]
                            
                        comp_label = "เอกสารครบถ้วน (9/9)" if comp["complete"] else f"เอกสารยังขาดไฟล์ที่: {', '.join(str(m) for m in comp['missing_indices'])}"
                        ai_res = {
                            "hospital_name": h_name,
                            "email": sender_email,
                            "date": email_date,
                            "mt_info": comp_label,
                            "folder_path": dest_dir,
                            "complete": comp["complete"],
                            "missing_indices": comp["missing_indices"],
                            "imported": True,
                            "last_updated": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        }
                        
                        # Save registry & Excel table
                        st.session_state.registry[h_name.lower()] = ai_res
                        append_to_excel(cfg.get("excel_path"), [ai_res])
                        
                        success_count += 1
                        log_usage(st.session_state.operator_name, f"นำเข้าข้อมูลด่วนสำเร็จ: {h_name}")
                    except Exception as e:
                        st.error(f"เกิดข้อผิดพลาดกับรายการ `{mail['subject'][:20]}`: {e}")
                        
                save_hospital_registry(st.session_state.registry)
                progress_bar.progress(1.0)
                progress_text.empty()
                st.session_state.show_success_modal = (success_count, total)
                st.session_state.fetched_emails = [] # Clear lists
                st.session_state.registry = load_hospital_registry()
                st.rerun()

# ================= MENU 2: ANALYZE DOCUMENT VALIDITY =================
elif st.session_state.active_menu == 2:
    st.subheader("🔍 เมนูที่ 2: วิเคราะห์และตรวจสอบความครบถ้วนของเอกสาร")
    st.write("เลือกโฟลเดอร์ที่เก็บไฟล์ เพื่อทำการวิเคราะห์ความครบถ้วนและถูกต้องของเอกสารทั้ง 9 ข้อ")
    
    # Resolve default path selection options
    workspace = st.session_state.workspace_dir
    dir_complete = os.path.join(workspace, "2. รพ ที่ verified แล้ว")
    dir_incomplete = os.path.join(workspace, "1. นำเข้าใหม่ยังไม่ได้เช็ค")
    
    # Gather folders inside complete/incomplete directories
    folder_options = []
    folder_paths = {}
    
    for parent_dir, label in [(dir_incomplete, "🆕 ยังไม่ได้เช็ค"), (dir_complete, "✅ Verified แล้ว")]:
        if os.path.exists(parent_dir):
            for d in sorted(os.listdir(parent_dir)):
                full_p = os.path.join(parent_dir, d)
                if os.path.isdir(full_p):
                    display_name = f"[{label}] {d}"
                    folder_options.append(display_name)
                    folder_paths[display_name] = full_p
                    
    # Custom browse path state
    if "custom_check_path" not in st.session_state:
        st.session_state.custom_check_path = ""
        
    if st.session_state.custom_check_path:
        display_name = f"📂 [โฟลเดอร์อื่นในเครื่อง] {os.path.basename(st.session_state.custom_check_path)}"
        if display_name not in folder_options:
            folder_options.insert(0, display_name)
            folder_paths[display_name] = st.session_state.custom_check_path
            
    # 1. Choose Folder Section
    st.markdown("#### 📁 1. เลือกโฟลเดอร์ที่เก็บไฟล์เอกสารโรงพยาบาล:")
    col_path_sel, col_path_browse = st.columns([3, 1])
    
    with col_path_sel:
        selected_display = st.selectbox(
            "โฟลเดอร์ที่เลือก:",
            options=folder_options,
            index=0 if folder_options else None,
            label_visibility="collapsed"
        )
    with col_path_browse:
        if st.button("📂 เลือกโฟลเดอร์อื่น...", use_container_width=True):
            custom_p = browse_directory()
            if custom_p:
                st.session_state.custom_check_path = custom_p
                st.rerun()
                
    active_path = folder_paths.get(selected_display) if selected_display else None
    
    if active_path:
        st.info(f"📍 เส้นทางโฟลเดอร์ปัจจุบัน: `{active_path}`")
        
        # 2. Check button
        st.markdown("#### 🔍 2. กดปุ่มตรวจสอบความครบถ้วนถูกต้อง:")
        run_check = st.button("🔍 ตรวจสอบความครบถ้วนถูกต้อง", type="primary", use_container_width=True)
        
        # We can store the check results in session state or run it on button click
        if "check_result" not in st.session_state:
            st.session_state.check_result = None
            st.session_state.check_result_path = ""
            
        if run_check or st.session_state.check_result_path != active_path:
            with st.spinner("กำลังวิเคราะห์ความถูกต้องและเวอร์ชันของเอกสาร..."):
                st.session_state.check_result = check_folder_completeness(active_path)
                st.session_state.check_result_path = active_path
                
        # 3. Display Results Section
        if st.session_state.check_result:
            comp = st.session_state.check_result
            missing = comp.get("missing_indices", [])
            is_complete = comp.get("complete", False)
            
            st.markdown("#### 📊 3. ผลการตรวจสอบเอกสาร:")
            
            # High-contrast banner showing what is missing
            if is_complete:
                st.markdown(f"""
                <div class="glass-card accent-border-green" style="background-color:#ECFDF5;">
                    <h3 style="color:#059669;margin:0;">✅ ครบถ้วนถูกต้องทั้งหมด (9/9 รายการ)</h3>
                    <p style="margin:4px 0 0 0;color:#1E293B;">เอกสารประกอบครบสมบูรณ์และเวอร์ชันถูกต้องตามเกณฑ์สภาเทคนิคการแพทย์</p>
                </div>
                """, unsafe_allow_html=True)
            else:
                missing_str = ", ".join(str(m) for m in missing)
                st.markdown(f"""
                <div class="glass-card accent-border-orange" style="background-color:#FEF2F2; border-left-color: #EF4444 !important;">
                    <h3 style="color:#DC2626;margin:0;">❌ ยังขาดเอกสารหรือเวอร์ชันไม่ถูกต้อง (ยังขาดรายการข้อ: {missing_str})</h3>
                    <p style="margin:4px 0 0 0;color:#1E293B;">กรุณาอัปโหลดเพิ่มเติมหรือทวงเอกสารหมายเลข: {missing_str}</p>
                </div>
                """, unsafe_allow_html=True)
                
            # Details report (9 items)
            file_descs = [
                "1. QM-LAB-001 คู่มือคุณภาพ",
                "2. F12 Checklist-MT2565 (V.4)",
                "3. HA Service profile",
                "4. F1 Lab_Profile_for_LA",
                "5. F7 Application_form_LA (5-2-67, ลงนาม)",
                "6. รายชื่อเจ้าหน้าที่",
                "7. LAB_SAFETY_Checklist (Version 2)",
                "8. QP-LAB-001 คู่มือห้องปฏิบัติการ",
                "9. FM-LAB-043 บัญชีรายชื่อเอกสาร"
            ]
            
            st.markdown("**🔍 รายละเอียดตรวจเช็คเวอร์ชันและความสมบูรณ์เชิงลึก:**")
            c_f1, c_f2 = st.columns(2)
            
            for file_idx in range(1, 10):
                target_col = c_f1 if file_idx <= 5 else c_f2
                status_text = comp["versions"][file_idx]
                is_ok = comp["versions_ok"][file_idx]
                
                icon = "🟢" if is_ok else "🔴"
                target_col.markdown(f"{icon} **{file_descs[file_idx - 1]}**<br><small style='color:#475569;'>{status_text}</small>", unsafe_allow_html=True)
                
            st.markdown("---")
            
            # Provide options to upload more files
            with st.expander("📂 อัปโหลดไฟล์เพิ่มเติมเข้าระบบ หรือ จัดการไฟล์"):
                col_up, col_zip = st.columns(2)
                with col_up:
                    uploaded_files = st.file_uploader(
                        "อัปโหลดไฟล์เอกสารเพิ่มเข้าโฟลเดอร์โรงพยาบาลนี้:", 
                        accept_multiple_files=True,
                        key="uploader_m2"
                    )
                    if uploaded_files:
                        for file in uploaded_files:
                            dest_file = os.path.join(active_path, file.name)
                            with open(dest_file, "wb") as f_out:
                                f_out.write(file.getbuffer())
                        st.success("อัปโหลดเอกสารเข้าโฟลเดอร์สำเร็จ! กรุณากดปุ่ม 'ตรวจสอบความครบถ้วนถูกต้อง' อีกครั้งเพื่อประเมินผลใหม่")
                        
                with col_zip:
                    st.write("ดาวน์โหลดไฟล์แนบสะสมทั้งหมด:")
                    zip_data = zip_folder(active_path)
                    st.download_button(
                        label="📥 ดาวน์โหลดไฟล์ทั้งหมดเป็น ZIP",
                        data=zip_data,
                        file_name=f"{os.path.basename(active_path)}_documents.zip",
                        mime="application/zip",
                        use_container_width=True
                    )
                    
            # Update status in Registry / Excel database
            # Try to identify which registry key matches this folder path
            h_registry_key = None
            h_name_found = ""
            for k, v in st.session_state.registry.items():
                if v.get("folder_path") == active_path or k.lower() in active_path.lower():
                    h_registry_key = k
                    h_name_found = v.get("hospital_name", k)
                    break
                    
            if h_registry_key:
                rec = st.session_state.registry[h_registry_key]
                st.markdown("---")
                
                # Let user Edit Profile Info optionally if they want
                with st.expander("📝 แก้ไขรายละเอียดข้อมูลประวัติโรงพยาบาล (ไม่บังคับ)"):
                    col_info1, col_info2 = st.columns(2)
                    e_hname = col_info1.text_input("ชื่อหน่วยงาน/โรงพยาบาล:", value=rec.get("hospital_name", h_name_found))
                    e_province = col_info2.text_input("จังหวัด:", value=rec.get("province", ""))
                    e_appointment = col_info1.text_input("ช่วงเวลาเสนอตรวจ:", value=rec.get("appointment", ""))
                    e_expiry = col_info2.text_input("วันหมดอายุใบอนุญาต:", value=rec.get("expiry_date", ""))
                    e_eval = col_info1.selectbox("ประเภทการตรวจ (LA/Re-LA):", ["LA", "Re-LA"], index=0 if rec.get("evaluation_type", "LA") == "LA" else 1)
                    e_contact = col_info2.text_input("ผู้ประสานงาน:", value=rec.get("contact_name", ""))
                    e_phone = col_info1.text_input("เบอร์โทรศัพท์:", value=rec.get("contact_phone", ""))
                    e_email = col_info2.text_input("อีเมล:", value=rec.get("email", ""))
                    
                col_save1, col_save2 = st.columns(2)
                
                with col_save1:
                    # Save local changes
                    if st.button("💾 บันทึกความคืบหน้า (Save Progress)", use_container_width=True):
                        rec["complete"] = is_complete
                        rec["missing_indices"] = missing
                        
                        # Apply optional field edits
                        rec["hospital_name"] = e_hname.strip()
                        rec["province"] = e_province.strip()
                        rec["appointment"] = e_appointment.strip()
                        rec["expiry_date"] = e_expiry.strip()
                        rec["evaluation_type"] = e_eval
                        rec["contact_name"] = e_contact.strip()
                        rec["contact_phone"] = e_phone.strip()
                        rec["email"] = e_email.strip()
                        
                        st.session_state.registry[h_registry_key] = rec
                        save_hospital_registry(st.session_state.registry)
                        st.success("บันทึกประวัติความคืบหน้าสำเร็จ!")
                        st.rerun()
                        
                with col_save2:
                    # Relocate folder to verified if complete and currently in folder 1
                    if is_complete and "1. นำเข้าใหม่ยังไม่ได้เช็ค" in active_path:
                        if st.button("✅ ยืนยันย้ายไปโฟลเดอร์ '2. รพ ที่ verified แล้ว'", type="primary", use_container_width=True):
                            email_date = rec.get("last_updated", datetime.datetime.now().strftime("%Y-%m-%d"))[:10]
                            new_path = get_hospital_folder_path(rec.get("hospital_name", h_name_found), email_date, True)
                            
                            # Move files
                            if os.path.exists(active_path) and active_path != new_path:
                                for item in os.listdir(active_path):
                                    shutil.copy2(os.path.join(active_path, item), os.path.join(new_path, item))
                                shutil.rmtree(active_path, ignore_errors=True)
                                
                            rec["folder_path"] = new_path
                            rec["complete"] = True
                            rec["missing_indices"] = []
                            st.session_state.registry[h_registry_key] = rec
                            
                            save_hospital_registry(st.session_state.registry)
                            # Update Excel
                            append_to_excel(cfg.get("excel_path"), [rec])
                            
                            st.session_state.custom_check_path = "" # Clear custom path
                            st.session_state.check_result = None
                            
                            st.success("ย้ายโฟลเดอร์ไปยังกลุ่ม Verified และบันทึกประวัติสำเร็จ!")
                            st.balloons()
                            st.rerun()
