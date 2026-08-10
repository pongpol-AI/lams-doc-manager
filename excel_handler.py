import openpyxl
import os
from copy import copy

def copy_cell_style(src_cell, dest_cell):
    """
    Copies the style (font, border, fill, number_format, alignment) 
    from src_cell to dest_cell to preserve existing document formatting.
    """
    if src_cell.font:
        dest_cell.font = copy(src_cell.font)
    if src_cell.border:
        dest_cell.border = copy(src_cell.border)
    if src_cell.fill:
        dest_cell.fill = copy(src_cell.fill)
    if src_cell.number_format:
        dest_cell.number_format = src_cell.number_format
    if src_cell.alignment:
        dest_cell.alignment = copy(src_cell.alignment)

def append_to_excel(file_path, items):
    """
    Dynamically maps items to existing Excel headers and appends them to the next row.
    Copies cell formatting (font, borders, fill, etc.) from the row above to match the sheet's look.
    If the destination file does not exist, copies a master template and clears its data rows to start fresh.
    """
    import shutil
    if not os.path.exists(file_path):
        # Create directories if they do not exist
        os.makedirs(os.path.dirname(os.path.abspath(file_path)), exist_ok=True)
        
        # Look for templates in parent directory and workspace root
        used_template = None
        templates = [
            os.path.join("ไฟล์ excel รวมรายชื่อโรงพยาบาล", "ตารางตรวจ LA สิงหาคม 69.xlsx"),
            "ตารางตรวจ LA สิงหาคม 69.xlsx",
            "July 2026 (5 ก.ค.69) 28 แห่ง.xlsx",
            "หน่วยงานที่ขอตรวจ.xlsx"
        ]
        for t in templates:
            if os.path.exists(t):
                used_template = t
                break
                
        if not used_template:
            # Fallback scan for any .xlsx
            for f in os.listdir("."):
                if f.endswith(".xlsx"):
                    used_template = f
                    break
                    
        if used_template:
            shutil.copy2(used_template, file_path)
            # Clear all data rows below detected header row to make it a fresh copy
            try:
                temp_wb = openpyxl.load_workbook(file_path)
                temp_sheet = temp_wb.active
                
                header_row = 1
                for r in range(1, 6):
                    c1_val = str(temp_sheet.cell(row=r, column=1).value or "").strip()
                    c2_val = str(temp_sheet.cell(row=r, column=2).value or "").strip()
                    if c1_val == "ลำดับ" or c2_val == "ลำดับ":
                        header_row = r
                        break
                        
                max_row = temp_sheet.max_row
                if max_row > header_row:
                    temp_sheet.delete_rows(header_row + 1, max_row - header_row)
                    
                temp_wb.save(file_path)
                temp_wb.close()
            except Exception:
                pass
        else:
            raise FileNotFoundError(f"ไม่พบไฟล์ Excel ปลายทางและไม่พบเทมเพลตสำหรับสร้างไฟล์ใหม่: {file_path}")
            
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        
        # 1. Detect header row (scan first 5 rows for "ลำดับ")
        header_row = 1
        for r in range(1, 6):
            c1_val = str(sheet.cell(row=r, column=1).value or "").strip()
            c2_val = str(sheet.cell(row=r, column=2).value or "").strip()
            if c1_val == "ลำดับ" or c2_val == "ลำดับ":
                header_row = r
                break
                
        # 2. Extract header list
        headers = []
        max_col = sheet.max_column
        for col_idx in range(1, max_col + 1):
            val = sheet.cell(row=header_row, column=col_idx).value
            headers.append(str(val or "").strip())
            
        print(f"Detected header row: {header_row}")
        print(f"Headers: {headers}")
        
        # 3. Find the actual last row with data (skipping empty styled/formatted cells)
        actual_max_row = header_row
        for r in range(sheet.max_row, header_row, -1):
            row_has_data = False
            for col_idx in range(1, max_col + 1):
                val = sheet.cell(row=r, column=col_idx).value
                if val is not None and str(val).strip() != "":
                    row_has_data = True
                    break
            if row_has_data:
                actual_max_row = r
                break
                
        # Find the last serial number (ลำดับ)
        last_id = 0
        if actual_max_row > header_row:
            for r in range(actual_max_row, header_row, -1):
                val = sheet.cell(row=r, column=1).value
                if isinstance(val, (int, float)):
                    last_id = int(val)
                    break
                elif str(val).isdigit():
                    last_id = int(val)
                    break
        
        # Start appending or updating rows
        current_row = actual_max_row + 1
        for item in items:
            h_name = str(item.get("hospital_name", "")).strip().lower()
            existing_row = None
            
            if h_name:
                # Find column index for hospital name
                h_col_idx = None
                for col_idx, header in enumerate(headers, 1):
                    if header in ["หน่วยงาน", "โรงพยาบาล"]:
                        h_col_idx = col_idx
                        break
                if h_col_idx:
                    for r in range(header_row + 1, actual_max_row + 1):
                        r_val = str(sheet.cell(row=r, column=h_col_idx).value or "").strip().lower()
                        if r_val and (r_val == h_name or h_name in r_val or r_val in h_name):
                            existing_row = r
                            break
                            
            target_row = existing_row if existing_row else current_row
            if not existing_row:
                last_id += 1
                row_serial = last_id
            else:
                row_serial = sheet.cell(row=existing_row, column=1).value or last_id
            
            # Map item dictionary keys to headers dynamically
            contact_name = item.get("contact_name", "")
            contact_phone = item.get("contact_phone", "")
            
            # Create a fallback full contact person string if separated fields are needed
            contact_person_full = f"{contact_name} {contact_phone}".strip()
            
            # Counter for duplicate "ผู้ตรวจ" columns
            assessor_counter = 0
            
            # Process each column in the sheet
            for col_idx, header in enumerate(headers, 1):
                new_cell = sheet.cell(row=target_row, column=col_idx)
                
                # Determine value based on header name
                val = None
                header_lower = header.lower()
                
                if header == "ลำดับ":
                    val = row_serial
                elif header in ["หน่วยงาน", "โรงพยาบาล"]:
                    val = item.get("hospital_name", "")
                elif header == "จังหวัด":
                    val = item.get("province", "")
                elif header == "Type":
                    val = item.get("evaluation_type", "")
                elif header == "ข้อมูล":
                    val = item.get("mt_info", "")
                elif header == "ผู้ประสานงาน":
                    val = contact_person_full
                elif header == "ผู้ติดต่อ":
                    val = contact_name
                elif header in ["เบอร์โทร", "โทร.", "โทร"]:
                    val = contact_phone
                elif header == "ที่อยู่":
                    val = item.get("address", "")
                elif header in ["ครบกำหนด", "วันครบกำหนด", "วันหมดอายุ", "หมดอายุ", "ครบกำหนดใบอนุญาต", "วันหมดอายุใบอนุญาต"]:
                    val = item.get("expiry_date", "")
                elif header in ["e-mail", "email", "อีเมล"]:
                    val = item.get("email", "")
                elif header == "ประเภทหน่วยงาน":
                    val = item.get("hospital_type", "")
                elif header == "การนัดหมาย":
                    val = item.get("appointment", "")
                elif header == "หัวหน้าทีม":
                    val = item.get("leader", "")
                elif "ผู้ตรวจ" in header:
                    if "1" in header:
                        val = item.get("assessor1", "")
                    elif "2" in header:
                        val = item.get("assessor2", "")
                    elif "3" in header:
                        val = item.get("assessor3", "")
                    elif "4" in header:
                        val = item.get("assessor4", "")
                    else:
                        # Handle duplicate "ผู้ตรวจ" headers (e.g. inหน่วยงานที่ขอตรวจ.xlsx)
                        assessor_counter += 1
                        if assessor_counter == 1:
                            val = item.get("assessor1", "")
                        elif assessor_counter == 2:
                            val = item.get("assessor2", "")
                        elif assessor_counter == 3:
                            val = item.get("assessor3", "")
                        elif assessor_counter == 4:
                            val = item.get("assessor4", "")
                else:
                    val = None
                    
                # If updating existing row, only set non-empty values unless explicitly provided
                if val is not None and str(val).strip() != "":
                    new_cell.value = val
                elif not existing_row:
                    new_cell.value = ""
                
                # Copy style from the row above if creating a new row
                if not existing_row and current_row > header_row + 1:
                    above_cell = sheet.cell(row=current_row - 1, column=col_idx)
                    copy_cell_style(above_cell, new_cell)
                    
            if not existing_row:
                current_row += 1
            
        wb.save(file_path)
        return len(items)
    except PermissionError:
        raise PermissionError(f"ไม่สามารถบันทึกไฟล์ได้ กรุณาปิดไฟล์ Excel '{os.path.basename(file_path)}' ก่อนทำรายการ")
    except Exception as e:
        raise Exception(f"เกิดข้อผิดพลาดในการเขียนไฟล์ Excel: {str(e)}")
